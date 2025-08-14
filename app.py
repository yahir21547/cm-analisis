# app_excel_azul_naranja.py
# Bloque AZUL (I2/Q2/A4:B22) intacto + Bloque NARANJA (I8/Q8/R3:S14), NEMA H3:H30.
# Coincidencia EXACTA para VLOOKUP; si no hay clave exacta => " ".

"""Tkinter application that mirrors the original Excel workbook.

Blue block
    • Input: HP (I2) and ambient °C (Q2).
    • Lookup table: A4:B22 to obtain the load fraction.

Orange block
    • Input: base HP (I8) and altitude FASL/MASL (Q8).
    • Lookup table: R3:S14 for altitude→load fraction mapping.

NEMA motor sizes are read from H3:H30.  All lookups require an exact
match; if the key is absent an empty string is shown.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

try:
    import openpyxl
except Exception:
    openpyxl = None
# --- Helper UI pieces originally from blue_orange_form.py ---

def create_row(parent, heading, fields):
    """Create a labelled row with one entry per field name."""
    row = ttk.Frame(parent)
    row.pack(fill="x", pady=5)
    ttk.Label(row, text=heading).grid(row=0, column=0, padx=5, sticky="w")
    for i, field in enumerate(fields):
        ttk.Label(row, text=field).grid(row=0, column=1 + 2 * i, padx=2, sticky="w")
        ttk.Entry(row, width=10).grid(row=0, column=2 + 2 * i, padx=2)


def build_basic_form(root):
    """Basic static form with blue/orange blocks; kept for reference."""
    root.title("Cuadros Azules y Naranjas")
    main = tk.Frame(root)
    main.pack(fill="both", expand=True)

    canvas = tk.Canvas(main)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar = ttk.Scrollbar(main, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)

    content = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=content, anchor="nw")

    def on_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    content.bind("<Configure>", on_configure)

    style = ttk.Style()
    style.configure("Blue.TLabelframe", background="#2882c7")
    style.configure("Blue.TLabelframe.Label", background="#2882c7", foreground="white", font=("Arial", 12, "bold"))
    style.configure("Orange.TLabelframe", background="#e98300")
    style.configure("Orange.TLabelframe.Label", background="#e98300", foreground="white", font=("Arial", 12, "bold"))

    blue = ttk.Labelframe(content, text="Cuadros AZULES", style="Blue.TLabelframe", padding=10)
    blue.pack(fill="x", pady=10)
    create_row(blue, "Base rating:", ["HP", "kW", "w"])
    create_row(blue, "50 Hz Rating:", ["Required HP", "NEMA HP"])
    create_row(blue, "Ambient Temperature:", ["Valor numérico", "Units (°C)"])
    create_row(blue, "Load %", ["%"])
    create_row(blue, "New Rating:", ["HP", "kW", "w", "NEMA HP"])
    create_row(blue, "Tolerancia (EC):", ["Required HP", "NEMA HP"])
    create_row(blue, "50 Hz Rating (otra sección):", ["Required HP", "NEMA HP"])
    create_row(blue, "Tolerancia (EC+50):", ["Required HP", "NEMA HP"])

    orange = ttk.Labelframe(content, text="Cuadros NARANJAS", style="Orange.TLabelframe", padding=10)
    orange.pack(fill="x", pady=10)
    create_row(orange, "Base rating:", ["HP", "kW", "w"])
    create_row(orange, "50 Hz Rating:", ["Required HP", "NEMA HP"])
    create_row(orange, "FASL (MASL):", ["Valor numérico", "Units (ft / m)"])
    create_row(orange, "Load %", ["%"])
    create_row(orange, "New Rating:", ["HP", "kW", "w", "NEMA HP"])
    create_row(orange, "Tolerancia (EC):", ["Required HP", "NEMA HP"])
    create_row(orange, "50 Hz Rating:", ["Required HP", "NEMA HP"])
    create_row(orange, "Tolerancia (EC+50):", ["Required HP", "NEMA HP"])

    def on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    canvas.bind_all("<MouseWheel>", on_mousewheel)


HP_PER_KW = 1.341
BLANK = " "

# === Config por defecto (solo precarga; puedes cambiarla en la UI) ===
DEFAULT_XLSX_PATH = r"C:\Users\MXYAGAR1\Downloads\piton\cm anailisis electrico\PMD NEMA V46 ADAPTED APPLICACION V1.xlsx"
DEFAULT_SHEET     = "cm electrico"

# Rangos fijos según tu hoja
# Azul -> A4:B22  (°C -> %)
A_COL, A_R0, B_COL, B_R1 = "A", 4, "B", 22
# Naranja -> R3:S14 (FASL/MASL -> %)
R_COL, R_R0, S_COL, S_R1 = "R", 3, "S", 14
# NEMA -> H3:H30 (pasos en HP)
NEMA_COL, NEMA_R0, NEMA_R1 = "H", 3, 30

FALLBACK_NEMA = [1,1.5,2,3,5,7.5,10,15,20,25,30,40,50,60,75,100,125,150,200,250,300,350,400,450,500,600,700,800]

# ---------- utilidades ----------
def to_kw(hp: float) -> float:
    return hp / HP_PER_KW

def to_watts(hp: float) -> float:
    return to_kw(hp) * 1000.0

def fmt(v, nd=2):
    if v == BLANK or v is None:
        return BLANK
    try:
        s = f"{float(v):.{nd}f}".rstrip("0").rstrip(".")
        return s
    except Exception:
        return str(v)

def pick_nema_hp(x, steps):
    if x is None or x <= 0:
        return BLANK
    for s in steps:
        try:
            if x <= float(s):
                return float(s)
        except Exception:
            pass
    try:
        top = float(steps[-1])
        return f">{int(top)} HP"
    except Exception:
        return ">800 HP"

def read_two_col_dict(path, sheet, col_key, r0, col_val, r1):
    """Lee un par de columnas numéricas a dict {key: val}."""
    if openpyxl is None:
        raise RuntimeError("Instala openpyxl: pip install openpyxl")
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {p}")
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        match = [s for s in wb.sheetnames if s.lower() == sheet.lower()]
        if not match:
            raise KeyError(f"No existe la hoja '{sheet}'. Hojas: {wb.sheetnames}")
        sheet = match[0]
    ws = wb[sheet]
    out = {}
    for r in range(r0, r1 + 1):
        k = ws[f"{col_key}{r}"].value
        v = ws[f"{col_val}{r}"].value
        if k is None or v is None:
            continue
        try:
            out[float(k)] = float(v)
        except Exception:
            pass
    wb.close()
    return out

def read_nema_steps(path, sheet, col="H", r0=3, r1=30):
    if openpyxl is None:
        raise RuntimeError("Instala openpyxl: pip install openpyxl")
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {p}")
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        match = [s for s in wb.sheetnames if s.lower() == sheet.lower()]
        if not match:
            raise KeyError(f"No existe la hoja '{sheet}'. Hojas: {wb.sheetnames}")
        sheet = match[0]
    ws = wb[sheet]
    vals = []
    for r in range(r0, r1 + 1):
        v = ws[f"{col}{r}"].value
        try:
            vals.append(float(v))
        except Exception:
            pass
    wb.close()
    vals = sorted(set(vals))
    return vals or FALLBACK_NEMA[:]

# ---------- app ----------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel → Tkinter (Bloque AZUL intacto + Bloque NARANJA)")
        self.geometry("1180x720")
        self.resizable(False, False)

        # tablas desde Excel
        self.tbl_blue   = {}   # A4:B22 (°C -> %)
        self.tbl_orange = {}   # R3:S14 (FASL/MASL -> %)
        self.nema_steps = FALLBACK_NEMA[:]

        self._build_ui()

        # precarga
        self.var_xlsx.set(DEFAULT_XLSX_PATH)
        self.var_sheet.set(DEFAULT_SHEET)
        self.load_from_excel(preload=True)

    # ---- UI helpers ----
    def _ro(self, parent, label, row, width=16, readonly=False):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=8, pady=6)
        v = tk.StringVar(value=BLANK)
        state = "readonly" if readonly else "normal"
        ttk.Entry(parent, width=width, textvariable=v, state=state).grid(row=row, column=1, padx=8, pady=6)
        return v

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        style = ttk.Style(self)
        style.configure("Blue.TLabelframe", background="#d9edf7")
        style.configure("Blue.TLabelframe.Label", background="#d9edf7")
        style.configure("Orange.TLabelframe", background="#ffe5b4")
        style.configure("Orange.TLabelframe.Label", background="#ffe5b4")

        # fuente de datos
        top = ttk.LabelFrame(self, text="Fuente (Excel)")
        top.place(x=10, y=10, width=1160, height=120)

        ttk.Label(top, text="Archivo:").grid(row=0, column=0, sticky="e", **pad)
        self.var_xlsx = tk.StringVar()
        ttk.Entry(top, width=80, textvariable=self.var_xlsx).grid(row=0, column=1, columnspan=3, **pad)
        ttk.Button(top, text="Cambiar…", command=self.pick_excel).grid(row=0, column=4, **pad)

        ttk.Label(top, text="Hoja:").grid(row=1, column=0, sticky="e", **pad)
        self.var_sheet = tk.StringVar()
        ttk.Entry(top, width=24, textvariable=self.var_sheet).grid(row=1, column=1, **pad)

        ttk.Button(top, text="Cargar tablas", command=self.load_from_excel).grid(row=1, column=2, **pad)
        ttk.Button(top, text="Limpiar todo", command=self.clear_all).grid(row=1, column=3, **pad)
        ttk.Button(top, text="Ver fórmulas", command=self.show_formulas).grid(row=1, column=4, **pad)

        self.lbl_status = ttk.Label(top, text="Tablas no cargadas")
        self.lbl_status.grid(row=2, column=0, columnspan=5, sticky="w", padx=10)

        # ===== BLOQUE AZUL (INTACTO) =====
        blue_in = ttk.LabelFrame(self, text="AZUL — Entradas", style="Blue.TLabelframe")
        blue_in.place(x=10, y=140, width=560, height=80)

        ttk.Label(blue_in, text="HP (I2):").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        self.i2_hp = tk.StringVar()
        ttk.Entry(blue_in, width=12, textvariable=self.i2_hp).grid(row=0, column=1, padx=8, pady=6)

        ttk.Label(blue_in, text="Ambient °C (Q2):").grid(row=0, column=2, sticky="w", padx=8, pady=6)
        self.q2_amb = tk.StringVar()
        ttk.Entry(blue_in, width=12, textvariable=self.q2_amb).grid(row=0, column=3, padx=8, pady=6)

        ttk.Button(blue_in, text="Calcular AZUL", command=self.calc_blue).grid(row=0, column=4, padx=8, pady=6)

        blue_out1 = ttk.LabelFrame(self, text="AZUL — Base rating (I2, I3, I4)", style="Blue.TLabelframe")
        blue_out1.place(x=10, y=230, width=360, height=150)
        self.b_i2  = self._ro(blue_out1, "HP (I2):", 0)
        self.b_i3  = self._ro(blue_out1, "kW (I3=I2/1.341):", 1)
        self.b_i4  = self._ro(blue_out1, "W (I4=kW*1000):", 2)

        blue_out2 = ttk.LabelFrame(self, text="AZUL — 50 Hz Rating (L2, L3)", style="Blue.TLabelframe")
        blue_out2.place(x=380, y=230, width=360, height=150)
        self.b_l2 = self._ro(blue_out2, "Required HP (L2=I2*1.15):", 0)
        self.b_l3 = self._ro(blue_out2, "NEMA HP (L3):", 1)

        blue_out3 = ttk.LabelFrame(self, text="AZUL — Ambient & Load (Q2, U2) / New Rating (Y2:Y4)", style="Blue.TLabelframe")
        blue_out3.place(x=10, y=390, width=730, height=180)
        self.b_q2  = self._ro(blue_out3, "Ambient (Q2):", 0)
        self.b_u2  = self._ro(blue_out3, "U2=VLOOKUP(A4:B22)/100:", 1)
        self.b_y2  = self._ro(blue_out3, "HP (Y2=I2/U2):", 2)
        self.b_y2kw= self._ro(blue_out3, "kW (Y3=Y2/1.341):", 3)
        self.b_y2w = self._ro(blue_out3, "W (Y4=kW*1000):", 4)
        self.b_y2n = self._ro(blue_out3, "NEMA HP (Y2):", 5)

        # ===== BLOQUE NARANJA =====
        orange_in = ttk.LabelFrame(self, text="NARANJA — Entradas", style="Orange.TLabelframe")
        orange_in.place(x=760, y=140, width=410, height=80)
        ttk.Label(orange_in, text="Base HP (I8):").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        self.i8_hp = tk.StringVar()
        ttk.Entry(orange_in, width=12, textvariable=self.i8_hp).grid(row=0, column=1, padx=8, pady=6)
        ttk.Label(orange_in, text="FASL/MASL (Q8):").grid(row=0, column=2, sticky="w", padx=8, pady=6)
        self.q8_fasl = tk.StringVar()
        ttk.Entry(orange_in, width=12, textvariable=self.q8_fasl).grid(row=0, column=3, padx=8, pady=6)
        ttk.Button(orange_in, text="Calcular NARANJA", command=self.calc_orange).grid(row=0, column=4, padx=8, pady=6)

        orange_base = ttk.LabelFrame(self, text="NARANJA — Base rating (I8,I9,I10) y 50 Hz (L8,L9)", style="Orange.TLabelframe")
        orange_base.place(x=760, y=230, width=410, height=150)
        self.o_i8  = self._ro(orange_base, "HP (I8):", 0)
        self.o_i9  = self._ro(orange_base, "kW (I9=I8/1.341):", 1)
        self.o_i10 = self._ro(orange_base, "W (I10=kW*1000):", 2)
        self.o_l8  = self._ro(orange_base, "Required HP (L8=I8*1.15):", 3)
        self.o_l9  = self._ro(orange_base, "NEMA HP (L9):", 4)

        orange_new = ttk.LabelFrame(self, text="NARANJA — Load (U8), New Rating (Y8) y Tolerancias", style="Orange.TLabelframe")
        orange_new.place(x=760, y=390, width=410, height=300)
        self.o_u8   = self._ro(orange_new, "U8=VLOOKUP(R3:S14)/100:", 0)
        self.o_y8   = self._ro(orange_new, "HP (Y8=I8/U8):", 1)
        self.o_y8kw = self._ro(orange_new, "kW (Y8/1.341):", 2)
        self.o_y8w  = self._ro(orange_new, "W ((Y8/1.341)*1000):", 3)
        self.o_y8n  = self._ro(orange_new, "NEMA HP (Y8):", 4)
        self.o_ab8  = self._ro(orange_new, "EC: AB8=Y8*0.94:", 5)
        self.o_ab9  = self._ro(orange_new, "NEMA(AB8):", 6)
        self.o_ae8  = self._ro(orange_new, "50Hz: AE8=Y8*1.15:", 7)
        self.o_ae9  = self._ro(orange_new, "NEMA(AE8):", 8)
        self.o_ah8  = self._ro(orange_new, "EC+50Hz: AH8=AE8*0.94:", 9)
        self.o_ah9  = self._ro(orange_new, "NEMA(AH8):", 10)

    # ---- Excel ----
    def pick_excel(self):
        path = filedialog.askopenfilename(
            title="Selecciona Excel",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm"), ("All files", "*.*")]
        )
        if path:
            self.var_xlsx.set(path)

    def load_from_excel(self, preload=False):
        try:
            p  = self.var_xlsx.get().strip()
            sh = (self.var_sheet.get().strip() or DEFAULT_SHEET)
            self.tbl_blue   = read_two_col_dict(p, sh, A_COL, A_R0, B_COL, B_R1)  # A4:B22
            self.tbl_orange = read_two_col_dict(p, sh, R_COL, R_R0, S_COL, S_R1)  # R3:S14
            self.nema_steps = read_nema_steps(p, sh, NEMA_COL, NEMA_R0, NEMA_R1)  # H3:H30
            self.lbl_status.config(text=f"Cargado: A4:B22({len(self.tbl_blue)}), R3:S14({len(self.tbl_orange)}), NEMA({len(self.nema_steps)})")
        except Exception as e:
            self.tbl_blue, self.tbl_orange = {}, {}
            self.nema_steps = FALLBACK_NEMA[:]
            self.lbl_status.config(text=f"No se pudo cargar: {e}")
            if not preload:
                messagebox.showerror("Error", str(e))

    # ---- Cálculos AZUL (idéntico a tus fórmulas) ----
    def calc_blue(self):
        # I2
        try:
            i2 = float(self.i2_hp.get().strip()) if self.i2_hp.get().strip() != "" else 0.0
        except ValueError:
            messagebox.showerror("Entrada inválida", "HP (I2) debe ser numérico.")
            return
        # Q2
        amb = None
        if self.q2_amb.get().strip() != "":
            try:
                amb = float(self.q2_amb.get().strip())
            except ValueError:
                messagebox.showerror("Entrada inválida", "Ambient °C (Q2) debe ser numérico.")
                return

        # Base: I3, I4
        if i2 == 0:
            self.b_i2.set(BLANK); self.b_i3.set(BLANK); self.b_i4.set(BLANK)
            base_ok = False
        else:
            self.b_i2.set(fmt(i2))
            self.b_i3.set(fmt(to_kw(i2)))
            self.b_i4.set(fmt(to_watts(i2)))
            base_ok = True

        # 50Hz L2, L3
        if not base_ok:
            self.b_l2.set(BLANK); self.b_l3.set(BLANK)
        else:
            l2 = i2 * 1.15
            self.b_l2.set(fmt(l2))
            self.b_l3.set(fmt(pick_nema_hp(l2, self.nema_steps)))

        # U2
        if amb is None or amb not in self.tbl_blue:
            self.b_q2.set(BLANK if amb is None else fmt(amb))
            self.b_u2.set(BLANK)
            u2 = None
        else:
            self.b_q2.set(fmt(amb))
            u2 = self.tbl_blue[amb] / 100.0
            self.b_u2.set(fmt(u2, nd=3))

        # Y2, kW, W, NEMA(Y2)
        if not base_ok or u2 is None:
            self.b_y2.set(BLANK); self.b_y2kw.set(BLANK); self.b_y2w.set(BLANK); self.b_y2n.set(BLANK)
        else:
            y2 = i2 / u2
            self.b_y2.set(fmt(y2))
            self.b_y2kw.set(fmt(to_kw(y2)))
            self.b_y2w.set(fmt(to_watts(y2)))
            self.b_y2n.set(fmt(pick_nema_hp(y2, self.nema_steps)))

    # ---- Cálculos NARANJA (idéntico a tus fórmulas) ----
    def calc_orange(self):
        # I8
        try:
            i8 = float(self.i8_hp.get().strip()) if self.i8_hp.get().strip() != "" else 0.0
        except ValueError:
            messagebox.showerror("Entrada inválida", "HP (I8) debe ser numérico.")
            return
        # Q8 (FASL/MASL)
        q8 = None
        if self.q8_fasl.get().strip() != "":
            try:
                q8 = float(self.q8_fasl.get().strip())
            except ValueError:
                messagebox.showerror("Entrada inválida", "FASL/MASL (Q8) debe ser numérico.")
                return

        # Base: I9, I10
        if i8 == 0:
            self.o_i8.set(BLANK); self.o_i9.set(BLANK); self.o_i10.set(BLANK)
            base_ok = False
        else:
            self.o_i8.set(fmt(i8))
            self.o_i9.set(fmt(to_kw(i8)))
            self.o_i10.set(fmt(to_watts(i8)))
            base_ok = True

        # 50Hz L8, L9
        if not base_ok:
            self.o_l8.set(BLANK); self.o_l9.set(BLANK)
        else:
            l8 = i8 * 1.15
            self.o_l8.set(fmt(l8))
            self.o_l9.set(fmt(pick_nema_hp(l8, self.nema_steps)))

        # U8 (lookup exacto en R3:S14)
        if q8 is None or q8 not in self.tbl_orange:
            self.o_u8.set(BLANK)
            u8 = None
        else:
            u8 = self.tbl_orange[q8] / 100.0
            self.o_u8.set(fmt(u8, nd=3))

        # Y8, kW, W, NEMA(Y8)
        if not base_ok or u8 is None:
            self.o_y8.set(BLANK); self.o_y8kw.set(BLANK); self.o_y8w.set(BLANK); self.o_y8n.set(BLANK)
            y8 = None
        else:
            y8 = i8 / u8
            self.o_y8.set(fmt(y8))
            self.o_y8kw.set(fmt(to_kw(y8)))
            self.o_y8w.set(fmt(to_watts(y8)))
            self.o_y8n.set(fmt(pick_nema_hp(y8, self.nema_steps)))

        # Tolerancia EC: AB8=Y8*0.94, AB9=NEMA(AB8)
        if y8 is None:
            self.o_ab8.set(BLANK); self.o_ab9.set(BLANK)
        else:
            ab8 = y8 * 0.94
            self.o_ab8.set(fmt(ab8))
            self.o_ab9.set(fmt(pick_nema_hp(ab8, self.nema_steps)))

        # 50Hz desde Y8: AE8=Y8*1.15, AE9=NEMA(AE8)
        if y8 is None:
            self.o_ae8.set(BLANK); self.o_ae9.set(BLANK)
            ae8 = None
        else:
            ae8 = y8 * 1.15
            self.o_ae8.set(fmt(ae8))
            self.o_ae9.set(fmt(pick_nema_hp(ae8, self.nema_steps)))

        # EC + 50Hz: AH8=AE8*0.94, AH9=NEMA(AH8)
        if ae8 is None:
            self.o_ah8.set(BLANK); self.o_ah9.set(BLANK)
        else:
            ah8 = ae8 * 0.94
            self.o_ah8.set(fmt(ah8))
            self.o_ah9.set(fmt(pick_nema_hp(ah8, self.nema_steps)))

    # ---- util ----
    def clear_all(self):
        for v in [
            # azul inputs
            self.i2_hp, self.q2_amb,
            # azul outs
            self.b_i2, self.b_i3, self.b_i4, self.b_l2, self.b_l3, self.b_q2, self.b_u2, self.b_y2, self.b_y2kw, self.b_y2w, self.b_y2n,
            # naranja inputs
            self.i8_hp, self.q8_fasl,
            # naranja outs
            self.o_i8, self.o_i9, self.o_i10, self.o_l8, self.o_l9, self.o_u8, self.o_y8, self.o_y8kw, self.o_y8w, self.o_y8n,
            self.o_ab8, self.o_ab9, self.o_ae8, self.o_ae9, self.o_ah8, self.o_ah9
        ]:
            v.set(BLANK)

    def show_formulas(self):
        text = (
            "Bloque AZUL:\n"
            "  I3 = I2 / 1.341\n"
            "  I4 = I3 * 1000\n"
            "  L2 = I2 * 1.15\n"
            "  L3 = NEMA(L2)\n"
            "  U2 = VLOOKUP(A4:B22)/100\n"
            "  Y2 = I2 / U2\n"
            "  Y3 = Y2 / 1.341\n"
            "  Y4 = Y3 * 1000\n"
            "\nBloque NARANJA:\n"
            "  I9 = I8 / 1.341\n"
            "  I10 = I9 * 1000\n"
            "  L8 = I8 * 1.15\n"
            "  L9 = NEMA(L8)\n"
            "  U8 = VLOOKUP(R3:S14)/100\n"
            "  Y8 = I8 / U8\n"
            "  AB8 = Y8 * 0.94\n"
            "  AE8 = Y8 * 1.15\n"
            "  AH8 = AE8 * 0.94\n"
        )
        messagebox.showinfo("Fórmulas", text)

    def pick_excel(self):
        path = filedialog.askopenfilename(
            title="Selecciona Excel",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm"), ("All files", "*.*")]
        )
        if path:
            self.var_xlsx.set(path)

# ===== Main =====
def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
